import sqlite3
from flask import jsonify
import openpyxl
import pandas as pd

def conect():
    DB_PATH = r"C:\Users\Usuario\Documents\DBHeromovil\VentasHeromovil.db"
    
    return DB_PATH

def empleados():
    DB_PATH = conect()
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""SELECT codigo, nombre
                        FROM usuarios
                        WHERE status = 1
                        AND puesto IN ('Ventas', 'Supervisor')
                        AND codigo NOT IN ('1003779', '1016312', '1020462')
                        ORDER BY nombre ASC;""")
    rows = cursor.fetchall()
    conn.close()

    return rows

def supervisor():
    DB_PATH = conect()
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    query = """
        SELECT codigo, nombre      
            FROM usuarios
            WHERE puesto = 'Supervisor'
			AND status = 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()

    return rows

def pagos(empleado_id):
    """
    Retorna la cantidad (conteo) de registros de la tabla 'pagos'
    que tengan el campo 'userlogin' igual al valor proporcionado.
    """    
    print(f"Empleado ID recibido: {empleado_id}")  # Debugging line
    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    if empleado_id and empleado_id.lower() != "none":
        query = """
        SELECT COUNT(*)
        FROM pagos
        WHERE SUBSTR(userlogin, 1, 7) = ?
        AND strftime('%Y-%m', dte) = strftime('%Y-%m', 'now', 'localtime');        
        """
        cur.execute(query, (empleado_id,))
    else:
        query = """
        SELECT COUNT(*)
        FROM pagos
        WHERE strftime('%Y-%m', dte) = strftime('%Y-%m', 'now', 'localtime');
        """
        cur.execute(query)

    resultado = cur.fetchone()
    conn.close()
    #Debugging lines
    """
    print(f"Consulta ejecutada: {query}")  # Debugging line
    print(f"Resultado de la consulta: {resultado}")  # Debugging line
    """

    if resultado is not None:
        return resultado[0]  # El primer (y único) valor es el conteo
    else:
        return 0

def get_recargas(empleado_id):
    """
    Retorna la cantidad (conteo) de registros de la tabla 'pagos'
    que tengan el campo 'userlogin' igual al valor proporcionado
    y el campo 'concepto_pago' igual a 'RECARGA-TRIVISION'.
    """    
    #print(f"Empleado ID recibido: {empleado_id}")  # Debugging line
    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    if empleado_id and empleado_id.lower() != "none":
        query = """
        SELECT SUM(monto) AS total_pago
        FROM pagos
        WHERE SUBSTR(userlogin, 1, 7) = ?
        AND concepto_pago = '876-PAGO/VENTA RECARGA-TRIVISION MONTO VARIA'
        AND strftime('%Y-%m', dte) = strftime('%Y-%m', 'now', 'localtime');        
        """
        cur.execute(query, (empleado_id,))
    else:
        query = """
        SELECT SUM(monto) AS total_pago
        FROM pagos
        WHERE concepto_pago = '876-PAGO/VENTA RECARGA-TRIVISION MONTO VARIA'
        AND strftime('%Y-%m', dte) = strftime('%Y-%m', 'now', 'localtime');
        """
        cur.execute(query)

    resultado = cur.fetchone()
    conn.close()
    #Debugging lines
    """ 
    print(f"Consulta ejecutada recargas: {query}")  # Debugging line
    print(f"Resultado de la consulta recargas: {resultado}")  # Debugging line
    """

    if resultado is not None:
        return resultado[0]  # El primer (y único) valor es el conteo
    else:
        return 0
    
def get_ventas(empleado_id=None):
    """
    Retorna la suma de total_ventas de la tabla 'ventas_detalle'.
    Puede ser filtrado por empleado_id.
    """
    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    query = """
    SELECT SUM(total_ventas) AS total_pav
    FROM ventas_detalle
    WHERE tipo_venta != 'Card'
    AND entity_code != 'EX332'
    AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime')
    """
    params = []

    if empleado_id and empleado_id.lower() != 'none' and empleado_id.strip() != '':
        query += " AND SUBSTR(usuario_creo_orden, 1, 7) = ?"
        params.append(empleado_id)

    cur.execute(query, params)

    resultado = cur.fetchone()
    conn.close()

    if resultado and resultado[0] is not None:
        return resultado[0]
    else:
        return 0

def get_rank_pav(start_date=None, end_date=None):
    """
    Retorna los top 10 usuarios (usuario_creo_orden) y su suma de ventas
    para un rango de fechas o para el mes actual, excluyendo 'Card'.
    """
    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    params = []
    date_filter = ""

    if start_date and end_date:
        date_filter = "fecha BETWEEN ? AND ?"
        params.extend([start_date, end_date])
    elif start_date:
        date_filter = "fecha = ?"
        params.append(start_date)
    elif end_date:
        date_filter = "fecha = ?"
        params.append(end_date)
    else:
        # Filtro por mes y año actual si no se proporcionan fechas
        date_filter = "strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime')"

    query = f"""
    SELECT usuario_creo_orden,
        SUM(total_ventas) AS total_ventas
    FROM ventas_detalle
    WHERE tipo_venta != 'Card'
    AND tipo_venta != 'CardEquipo'
    AND tipo_venta != 'InternetCard'
    AND entity_code  != 'EX332'
    AND {date_filter}
    GROUP BY usuario_creo_orden
    ORDER BY total_ventas DESC
    LIMIT 10;
    """
    cur.execute(query, tuple(params))
    filas = cur.fetchall()

    conn.close()

    return filas

def get_rank_pav_cc():
    """
    Retorna los top 10 usuarios (usuario_creo_orden) y su suma de ventas
    para el mes/año actual, excluyendo 'entidad 44066'.
    """
    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    query = """
    SELECT usuario_creo_orden,
        SUM(total_ventas) AS total_ventas
    FROM ventas_detalle
    WHERE tipo_venta != 'Card'
    AND entity_code  != '44066'
    AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime')
    AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime')
    GROUP BY usuario_creo_orden
    ORDER BY total_ventas DESC
    LIMIT 10;
    """
    cur.execute(query)
    # fetchall para todas las filas
    filas = cur.fetchall()

    conn.close()

    # 'filas' será una lista de tuplas: [(usuario_creo_orden, total_ventas), ...]
    return filas

def get_recent_activity():
    """
    Obtiene los 3 tipos de venta con más volumen, funcion modificada poara mostrar el dia, y excluir 'Card'.
    1. Si hoy es domingo, obtiene los datos del sábado.
    2. Si no, obtiene los datos de hoy.
    mantener este formato para la compatibilidad con el front-end.
    """
    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    try:        
        query = """
        SELECT usuario_creo_orden, tipo_venta, SUM(total_ventas) AS cantidad
        FROM ventas_detalle
        WHERE date(fecha) = CASE		
            WHEN strftime('%w', 'now', 'localtime') = '0' 
                THEN date('now', '-1 day', 'localtime')   -- si es domingo, usar el sábado
            ELSE date('now', 'localtime')                  -- si no, usar hoy
        END
        AND tipo_venta != 'Card'
        AND entity_code != 'EX332'
        GROUP BY usuario_creo_orden, tipo_venta
        ORDER BY cantidad DESC
        LIMIT 7;
        """
        cur.execute(query)
        
        data = cur.fetchall()

        cur.close()

        return data
    except Exception as e:
        print(f"Error en get_recent_activity: {e}")
        return []

def get_sales_overview():
    """
    Obtiene los 5 tipos de venta más importantes (excluyendo 'Card') del mes actual
    y calcula su porcentaje respecto al total (sin 'Card').
    """
    conn = sqlite3.connect(conect())
    cur = conn.cursor()
    try:
        # Obtener el total de ventas excluyendo 'Card' para el mes actual
        query_total = """
        SELECT SUM(total_ventas) as total
        FROM ventas_detalle
        WHERE tipo_venta != 'Card' 
        AND entity_code != 'EX332'
        AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime');
        """
        cur.execute(query_total)
        total_sales_result = cur.fetchone()
        total_sales = total_sales_result[0] if total_sales_result else 0

        if total_sales is None or total_sales == 0:
            cur.close()
            conn.close()
            return {"labels": [], "data": []}

        # Obtener el top 5 de ventas (sin 'Card') para el mes actual
        query_top5 = """
        SELECT tipo_venta, SUM(total_ventas) AS total
            FROM ventas_detalle
            WHERE tipo_venta IN (
                'Flex/Max',
                'Migraciones',
                'Internet',
                'Reemplazo Aumento',
                'Aumentos Pospago',
                'Aumentos Internet'
            )
            AND entity_code != 'EX332'
            AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime')
            GROUP BY tipo_venta
            ORDER BY total DESC;
        """
        cur.execute(query_top5)
        top5_sales = cur.fetchall()
        cur.close()
        conn.close()

        labels = []
        data = []
        for row in top5_sales:
            labels.append(row[0])
            percentage = (row[1] / total_sales) * 100
            data.append(round(percentage, 2))

        return {"labels": labels, "data": data}
    except Exception as e:
        print(f"Error en get_sales_overview: {e}")
        # Asegurarse de cerrar la conexión si hay un error
        if 'cur' in locals() and cur:
            cur.close()
        if 'conn' in locals() and conn:
            conn.close()
        return {"labels": [], "data": []}

# Guarda el DataFrame 'filtrado' en la base de datos SQLite en tabla transacciones
def guardar_filtrado_en_db(filtrado):
    # Reemplaza NaN por None
    filtrado = filtrado.where(pd.notnull(filtrado), None)

    # Conexión
    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    TABLE_NAME = "transacciones"

    # 1️ Crear tabla con restricción UNIQUE
    create_table_query = f"""
    CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_transaccion TEXT NOT NULL,
        fecha_digitacion_orden DATE,
        fecha_termino_orden DATE,
        estado_transaccion TEXT,
        usuario_creo_orden TEXT NOT NULL,
        entity_code TEXT,
        subcanal INTEGER,
        tipo_actividad TEXT,
        razon_servicio TEXT NOT NULL,
        telefono TEXT NOT NULL,
        imei TEXT NOT NULL,
        nom_plan TEXT,
        grupo_activacion_orden TEXT,
        grupo_activacion_anterior TEXT,
        UNIQUE (id_transaccion, usuario_creo_orden, razon_servicio, telefono)
    );
    """
    cur.execute(create_table_query)
    

    # 2️ Insertar o actualizar (Usa la versión corregida que te di antes)
    insert_query = f"""
    INSERT INTO {TABLE_NAME} (
        id_transaccion, fecha_digitacion_orden, fecha_termino_orden,
        estado_transaccion, usuario_creo_orden, entity_code, subcanal, tipo_actividad,
        razon_servicio, telefono, imei, nom_plan, grupo_activacion_orden, grupo_activacion_anterior
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ON CONFLICT(id_transaccion, usuario_creo_orden, razon_servicio, telefono)
    DO UPDATE SET
        estado_transaccion = excluded.estado_transaccion,
        fecha_termino_orden = excluded.fecha_termino_orden
    WHERE 
        (transacciones.estado_transaccion = 'Abierta' AND excluded.estado_transaccion = 'Terminada')
        OR
        (transacciones.estado_transaccion = 'Abierta' AND excluded.estado_transaccion = 'Cancelada')
        OR
        (transacciones.estado_transaccion != excluded.estado_transaccion);
    """ #  lógica corregida de solicitud anterior

    # 3️ Ejecutar en bloque (Ahora sin las conversiones str())
    data = [
        (
            row["id_transaccion"], # Ya es string o None
            str(row["fecha_digitacion_orden"]) if pd.notnull(row["fecha_digitacion_orden"]) else None,
            str(row["fecha_termino_orden"]) if pd.notnull(row["fecha_termino_orden"]) else None,
            row["estado_transaccion"],
            row["usuario_creo_orden"],
            row["entity_code"],
            row["subcanal"], # Ya es Int64 o None
            row["tipo_actividad"],
            row["razon_servicio"],
            row["telefono"], # Ya es string o None
            row["imei"], # Ya es string o None
            row["nom_plan"],
            row["grupo_activacion_orden"],
            row["grupo_activacion_anterior"],
        )
        for _, row in filtrado.iterrows()
    ] 

    cur.executemany(insert_query, data)

    conn.commit()
    conn.close()

# Guarda el los datos procedados en la base de datos SQLite ventas_detalle
def guardar_ventas_detalle_en_db(df_ventas):

    conn = sqlite3.connect(conect())
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ventas_detalle (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_code TEXT,
            subcanal INTEGER,
            fecha DATE,
            supervisor TEXT,
            usuario_creo_orden TEXT,
            tipo_venta TEXT,
            Grupo INTEGER,
            Grupo_Anterior INTEGER,
            total_ventas INTEGER,  
            Comision_100 INTEGER,
            Comision_75 INTEGER        
        );
    """
    )

    # OPCIONAL: borrar todo antes de insertar
    cur.execute("DELETE FROM ventas_detalle")
    conn.commit()

    insert_sql = """
        INSERT INTO ventas_detalle (
            entity_code,
            subcanal,
            fecha,
            supervisor,
            usuario_creo_orden,
            tipo_venta,
            Grupo,
            Grupo_Anterior,
            total_ventas,
            Comision_100,
            Comision_75            
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) 
        """

    for _, row in df_ventas.iterrows():
        values = (
            row["entity_code"],
            row["subcanal"],
            str(row["fecha"]),  # str() si viene como datetime
            row["supervisor"],
            row["usuario_creo_orden"],
            row["tipo_venta"],
            int(row["Grupo"]),
            int(row["Grupo_Anterior"]),
            int(row["total_ventas"]),
            int(row["Comision_100"]),
            int(row["Comision_75"]),
        )
        cur.execute(insert_sql, values)

    conn.commit()
    conn.close()

# Guarda los pagos desde un archivo Excel a la tabla 'pagos'
def insertar_pagos(df_pagos):
    try:
        #print("Datos Antes Insert:")
        #print(df_pagos.head())

        # Conectar a la base de datos SQLite
        conn = sqlite3.connect(conect())
        cursor = conn.cursor()

        # 1️ Crear tabla si no existe
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS pagos (
            id_tipo_compania INTEGER,
            compania TEXT,
            ent_code TEXT,
            estado TEXT,
            entity_name	TEXT,
            id_canal INTEGER,
            id_subcanal	INTEGER,
            dte	TEXT,
            monto REAL,
            custcode TEXT,
            id_cuenta TEXT,
            cn TEXT,
            fn TEXT,
            cachknum TEXT,
            userlogin TEXT,
            caja TEXT,
            concepto_pago TEXT,
            transaction_id INTEGER,
            fp_efectivo REAL,
            fp_tarjeta REAL,
            fp_cheque REAL,
            fp_otras REAL,
            tel_contacto TEXT,
            tel_contacto2 TEXT,
            UNIQUE(ent_code, transaction_id, custcode, dte, monto)
        );
        """)

        # 2️ Preparar datos desde el DataFrame
        df_pagos = df_pagos.where(pd.notnull(df_pagos), None)
        datos_a_insertar = [tuple(x) for x in df_pagos.to_numpy()]
        print("Ejemplo de fila a insertar:")
        print(datos_a_insertar[0])
        print(f"Cantidad de registros a insertar: {len(datos_a_insertar)}")        

        print(f"Filas afectadas realmente: {cursor.rowcount}")

        # 4️ Insertar datos (evita duplicados con INSERT OR IGNORE)
        cursor.executemany("""
            INSERT OR IGNORE INTO pagos (
                id_tipo_compania, compania, ent_code, estado, entity_name,
                id_canal, id_subcanal, dte, monto, custcode, id_cuenta, cn, fn,
                cachknum, userlogin, caja, concepto_pago, transaction_id,
                fp_efectivo, fp_tarjeta, fp_cheque, fp_otras, tel_contacto, tel_contacto2
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, datos_a_insertar)

        # 5️ Guardar cambios
        conn.commit()
        conn.close()
        print(f"{len(datos_a_insertar)} registros procesados correctamente (sin duplicados).")
        return True

    except Exception as e:        
        print(f"Hubo un problema al insertar los datos: {e}")
        raise e
       
#Guarda los objetivos desde un archivo Excel a la tabla 'objetivos'
def insertar_objetivos(df_objetivos):

    try:
        # Conectar a la base de datos SQLite
        conn = sqlite3.connect(conect())
        cursor = conn.cursor()

        # Leer el archivo Excel
        wb = openpyxl.load_workbook(df_objetivos)
        if wb.sheetnames:
            hoja = wb.active
        else:
            raise ValueError("El archivo Excel no contiene hojas.")

        # Iterar sobre las filas del archivo Excel e insertar en la tabla
        for fila in hoja.iter_rows(min_row=2, values_only=True): # type: ignore
            cursor.execute("""
                INSERT INTO objetivos (id_empleado, fecha, sim_card_prepago, flex_max, fijos_hfc_dth, internet_pospago,
                                    migraciones_pospago_net, fidepuntos_pospago, fidepuntos_up, reemplazo_pospago,
                                    reemplazo_up, fide_reemp_internet_up, aumentos_plan_pos_net, recargas)
                VALUES (?, ?,?,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, fila)

        # Confirmar la transacción y cerrar la conexión
        conn.commit()
        conn.close()
        print("Objetivos insertados correctamente")

    except Exception as e:
        print(f"Hubo un problema al insertar los objetivos: {e}")

def generar_ventas():
    """
    Genera un reporte de ventas basado en los datos almacenados en la base de datos.
    Agrupa las ventas por diferentes categorías y guarda el resultado en la base de datos.
    """
    try:
        conn = sqlite3.connect(conect()) 

        query = """
        SELECT *FROM transacciones
	        WHERE estado_transaccion != 'Cancelada'
        """
        query2 = """
        SELECT codigo, supervisor FROM usuarios
        """
        query3 = """
        SELECT *FROM incentivosEmpleados
        """
        data = pd.read_sql(query, conn)
        data_supers = pd.read_sql(query2, conn)
        data_incentivos = pd.read_sql(query3, conn)        

        # Convertir la columna 'fecha_digitacion_orden' a datetime y establecer la hora a las 00:00:00
        data["fecha_digitacion_orden"] = pd.to_datetime(
            data["fecha_digitacion_orden"], errors="coerce"
        ).dt.normalize() 

        # Convertir los datos en un DataFrame
        df_group = data.copy()
        df_group_Intermet = data.copy()
        df_group_Card = data.copy()

        # Definir los grupos de ventas y sus condiciones
        grupos_ventas = {
            "Flex/Max": (
                df_group["razon_servicio"].isin(
                    [
                        "1425 - ACTIVACIONES DE LINEAS - SIN EQUIPOS (DEALERS)",
                        "1409 - PORT SIN NUMERO TEMPORERO/SIN EQUIPO (CANAL PRESE)",
                        "1126 - ACTIVACIONES DE LÍNEAS - CON EQUIPOS (EFECTIVO)",
                        "1301 - ACTIVACIONES DE LINEAS - CON EQUIPOS",
                        "1302 - PORT IN CON NUMERO TEMPORERO/CON EQUIPO",
                        "1349 - PORT CON NUMERO TEMPORERO/SIN EQUIPO (CANAL PRESE)",
                        "1402 - ACTIVACIONES LÍNEAS-SIN EQUIPOS (CANAL PRESENCIAL)",
                    ]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Migraciones": (
                df_group["razon_servicio"].isin(
                    [
                        "1468 - AUMENTO DE PRODUCTO CON VENTA DE EQUIPO",                        
                        "1316 - AUMENTO DE PRODUCTO PREPAGO-POSTPAGO",                        
                    ]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Fidepuntos Pospago": (
                df_group["razon_servicio"].isin(
                    ["1378 - USO DE FIDEPUNTOS (CAMBIA TU MOVIL)"]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Fidepuntos Aumento": (
                df_group["razon_servicio"].isin(
                    [
                     "5001 - AUMENTO DE PLAN CON VENTA EQUIPO POR FIDE",
                     "5201 - AUMENTO DE PRODUCTO CON VENTA EQUIPO POR FIDE",
                     ]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Fidepuntos Disminucion": (
                df_group["razon_servicio"].isin(
                    ["5031 - DISMINUCIÓN DE PLAN CON VENTA EQUIPO POR FIDE"]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Reemplazo Pospago": (
                df_group["razon_servicio"].isin(["1324 - VENTA DE EQUIPOS/REEMPLAZO"])
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Reemplazo Aumento": (
                df_group["razon_servicio"].isin(
                    [
                     "5000 - AUMENTO DE PLAN CON VENTA EQUIPO POR REEMPLAZO",
                     "5200 - AUMENTO DE PRODUCTO CON VENTA EQUIPO POR REEMPLAZO",
                     ]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Reemplazo Disminucion": (
                df_group["razon_servicio"].isin(
                    ["5030 - DISMINUCIÓN DE PLAN CON VENTA EQUIPO POR REEMPLAZO"]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
            "Aumentos Pospago": (
                df_group["razon_servicio"].isin(
                    [
                    "1322 - AUMENTO DE PLAN / PAQUETE",
                    "1317 - AUMENTO DE PRODUCTO POSTPAGO-POSTPAGO",
                    ]
                    )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"883 - GRUPO 0|884 - GRUPO 1|885 - GRUPO 2|886 - GRUPO 3|887 - GRUPO 4|"
                    r"888 - GRUPO 5|889 - GRUPO 6|890 - GRUPO 7|891 - GRUPO 8|893 - GRUPO 9|"
                    r"894 - GRUPO 10|895 - GRUPO 11|912 - GRUPO 12",
                    case=False,
                    na=False,
                )
            ),
        }

        grupos_ventas_internet = {
            "Internet": (
                df_group_Intermet["razon_servicio"].isin(
                    [
                        "1301 - ACTIVACIONES DE LINEAS - CON EQUIPOS",
                        "1402 - ACTIVACIONES LÍNEAS-SIN EQUIPOS (CANAL PRESENCIAL)",
                    ]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "InternetCard": (
                df_group_Intermet["razon_servicio"].isin(
                    [
                        "1301 - ACTIVACIONES DE LINEAS - CON EQUIPOS",
                        "1402 - ACTIVACIONES LÍNEAS-SIN EQUIPOS (CANAL PRESENCIAL)",
                    ]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"482 - GRUPO NET 0 - 12 MESES",
                    case=False,
                    na=False,
                )
            ),
            "Migracion Internet": (
                df_group_Intermet["razon_servicio"].isin(
                    [
                        "1468 - AUMENTO DE PRODUCTO CON VENTA DE EQUIPO",
                        "1316 - AUMENTO DE PRODUCTO PREPAGO-POSTPAGO",
                    ]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "Reemplazo Internet": (
                df_group_Intermet["razon_servicio"].isin(
                    ["1324 - VENTA DE EQUIPOS/REEMPLAZO"]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "Reemplazo InternetAum": (
                df_group_Intermet["razon_servicio"].isin(
                    ["5000 - AUMENTO DE PLAN CON VENTA EQUIPO POR REEMPLAZO"]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "Reemplazo InternetDism": (
                df_group_Intermet["razon_servicio"].isin(
                    ["5030 - DISMINUCIÓN DE PLAN CON VENTA EQUIPO POR REEMPLAZO"]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "Fidepuntos Internet": (
                df_group_Intermet["razon_servicio"].isin(
                    ["1378 - USO DE FIDEPUNTOS (CAMBIA TU MOVIL)"]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "Fidepuntos InternetAum": (
                df_group_Intermet["razon_servicio"].isin(
                    ["5001 - AUMENTO DE PLAN CON VENTA EQUIPO POR FIDE"]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GR UPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "Fidepuntos InternetDism": (
                df_group_Intermet["razon_servicio"].isin(
                    ["5031 - DISMINUCIÓN DE PLAN CON VENTA EQUIPO POR FIDE"]
                )
                & df_group_Intermet["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
            "Aumentos Internet": (
                df_group_Intermet["razon_servicio"].isin(
                    ["1322 - AUMENTO DE PLAN / PAQUETE"]
                )
                & df_group["grupo_activacion_orden"].str.contains(
                    r"464 - GRUPO NET 1 - 18 MESES|465 - GRUPO NET 2 - 18 MESES|466 - GRUPO NET 3 - 18 MESES|"
                    r"467 - GRUPO NET 4 - 18 MESES|471 - GRUPO NET 5 - 18 MESES|802 - GRUPO NET 6 - 18 MESES|"
                    r"803 - GRUPO NET 7 - 18 MESES|",
                    case=False,
                    na=False,
                )
            ),
        }

        grupos_ventas_card = {
            "Card": (
                df_group_Card["razon_servicio"].isin(
                    [
                        "1425 - ACTIVACIONES DE LINEAS - SIN EQUIPOS (DEALERS)",
                        "1409 - PORT SIN NUMERO TEMPORERO/SIN EQUIPO (CANAL PRESE)",                        
                        "1349 - PORT CON NUMERO TEMPORERO/SIN EQUIPO (CANAL PRESE)",
                        "1402 - ACTIVACIONES LÍNEAS-SIN EQUIPOS (CANAL PRESENCIAL)",
                    ]
                )
                & df_group_Card["grupo_activacion_orden"].str.contains(
                    r"322 - GRUPO CARD SIM ONLY|4 - CARD",
                    case=False,
                    na=False,
                )
            ),
            "CardEquipo": (
                df_group_Card["razon_servicio"].isin(
                    [
                        "1126 - ACTIVACIONES DE LÍNEAS - CON EQUIPOS (EFECTIVO)",
                        "1301 - ACTIVACIONES DE LINEAS - CON EQUIPOS",
                        "1302 - PORT IN CON NUMERO TEMPORERO/CON EQUIPO",                        
                    ]
                )
                & df_group_Card["grupo_activacion_orden"].str.contains(
                    r"322 - GRUPO CARD SIM ONLY|4 - CARD",
                    case=False,
                    na=False,
                )
            ),
        }

        # Agrupación card
        df_group_Card["tipo_venta"] = None
        for tipo, condicion in grupos_ventas_card.items():
            df_group_Card.loc[condicion, "tipo_venta"] = tipo

        # Extraer Grupo a partir de dos patrones
        df_group_Card["Grupo1"] = df_group_Card["grupo_activacion_orden"].str.extract(
            r"(\d+) - GRUPO CARD SIM ONLY", expand=False
        )
        df_group_Card["Grupo2"] = df_group_Card["grupo_activacion_orden"].str.extract(
            r"(\d+) - CARD", expand=False
        )

        df_group_Card["Grupo"] = df_group_Card["Grupo1"].fillna(df_group_Card["Grupo2"])
        df_group_Card.drop(["Grupo1", "Grupo2"], axis=1, inplace=True)
        # fin card

        # Agrupación Flex/Max
        df_group["tipo_venta"] = None
        for tipo, condicion in grupos_ventas.items():
            df_group.loc[condicion, "tipo_venta"] = tipo

        # Extraer el número del grupo para agruparlo
        df_group["Grupo"] = (
            df_group["grupo_activacion_orden"]
            .str.extract(r"GRUPO (\d+)", expand=True)
            .bfill(axis=1)[0]        
        )
        

        # Extraer el número del grupo anterior para agruparlo
        df_group["Grupo_Anterior"] = (
            df_group["grupo_activacion_anterior"]
            .str.extract(r"GRUPO (\d+)", expand=True)
            .bfill(axis=1)[0]
        )
        # fin flex/mas

        # Agrupación Internet
        df_group_Intermet["tipo_venta"] = None
        for tipo, condicion in grupos_ventas_internet.items():
            df_group_Intermet.loc[condicion, "tipo_venta"] = tipo

        # Extraer el número del grupo para agruparlo
        df_group_Intermet["Grupo"] = (
            df_group_Intermet["grupo_activacion_orden"]
            .str.extract(r"GRUPO NET (\d+)", expand=True)
            .bfill(axis=1)[0]
        )

        # Extraer el número del grupo anterior para agruparlo
        df_group_Intermet["Grupo_Anterior"] = (
            df_group_Intermet["grupo_activacion_anterior"]
            .str.extract(r"GRUPO NET (\d+)", expand=True)
            .bfill(axis=1)[0]
        )

        df_group_Intermet["Grupo_Anterior"] = df_group_Intermet["Grupo_Anterior"].fillna(0)
        # Fin Agrupacion Internet

        # Rellenar NaN en "Grupo Anterior" con 0
        df_group["Grupo_Anterior"] = df_group["Grupo_Anterior"].fillna(0)

        # Generar conteo
        ventas_validas = df_group.dropna(subset=["tipo_venta", "Grupo"], how="any")
        conteo_ventas = (
            ventas_validas.groupby(
                [
                    "entity_code",
                    "subcanal",
                    "fecha_digitacion_orden",                    
                    "usuario_creo_orden",                    
                    "tipo_venta",
                    "Grupo",
                    "Grupo_Anterior",
                ]
            )
            .size()
            .reset_index()
        )

        # Internet
        ventas_validas_internet = df_group_Intermet.dropna(
            subset=["tipo_venta", "Grupo"], how="any"
        )
        conteo_ventas_internet = (
            ventas_validas_internet.groupby(
                [
                    "entity_code",
                    "subcanal",
                    "fecha_digitacion_orden",                                        
                    "usuario_creo_orden",
                    "tipo_venta",
                    "Grupo",
                    "Grupo_Anterior",
                ]
            )
            .size()
            .reset_index()
        )

        # card
        ventas_validas_card = df_group_Card.dropna(subset=["tipo_venta"], how="any")
        conteo_ventas_card = (
            ventas_validas_card.groupby(
                [
                    "entity_code",
                    "subcanal",
                    "fecha_digitacion_orden",                                       
                    "usuario_creo_orden",
                    "tipo_venta",
                    "Grupo",
                ]
            )
            .size()
            .reset_index()
        )

        # Renombrar la columna de conteo
        conteo_ventas.rename(columns={0: "total_ventas"}, inplace=True)
        conteo_ventas_internet.rename(columns={0: "total_ventas"}, inplace=True)
        conteo_ventas_card.rename(columns={0: "total_ventas"}, inplace=True)

        # Verificar si los DataFrames están vacíos
        if conteo_ventas.empty:
            raise ValueError("No se encontraron ventas válidas.")

        if conteo_ventas_internet.empty:
            raise ValueError("No se encontraron ventas válidas.")

        if conteo_ventas_card.empty:
            raise ValueError("No se encontraron ventas válidas.")

        # Preparar los resultados individuales
        resultado = conteo_ventas.reset_index()
        resultado_card = conteo_ventas_card.reset_index()
        resultado_internet = conteo_ventas_internet.reset_index()

        # Concatenar todos los resultados
        reultado_final = pd.concat(
            [resultado, resultado_internet, resultado_card], ignore_index=True
        )
        reultado_final.rename(columns={"fecha_digitacion_orden": "fecha"}, inplace=True)

        # Rellenar valores nulos en Grupo_Anterior
        reultado_final["Grupo_Anterior"] = reultado_final["Grupo_Anterior"].fillna(0)        

        # Extraer código de usuario para asignar supervisores
        reultado_final["codigo"] = reultado_final["usuario_creo_orden"].str[:7]
        data_supers["codigo"] = data_supers['codigo'].astype(str)        

        # Unir con datos de supervisores
        df_combinado = pd.merge(reultado_final, data_supers, on="codigo", how="left")
        df_combinado = df_combinado.fillna(0)

        # Convertir supervisor a entero sin decimales
        df_combinado["supervisor"] = pd.to_numeric(df_combinado["supervisor"], errors='coerce')
        df_combinado["supervisor"] = df_combinado["supervisor"].fillna(0).astype(int).astype(str)

        # Lista de tipos de venta para los que Grupo_Anterior debe ser 0
        tipos_venta_grupo_anterior_cero = [
            "Reemplazo InternetAum",
            "Fidepuntos InternetDism",
            "Fidepuntos InternetAum",
            "Reemplazo InternetDism",
            "Reemplazo Disminucion",
            "Reemplazo Aumento",
            "Fidepuntos Disminucion",
            "Fidepuntos Aumento",
            "Migraciones",
        ]

        # Establecer Grupo_Anterior a 0 para los tipos de venta especificados
        df_combinado.loc[df_combinado['tipo_venta'].isin(tipos_venta_grupo_anterior_cero), 'Grupo_Anterior'] = 0

        # Reorganizar columnas
        columnas = df_combinado.columns.tolist()
        columnas.insert(
            columnas.index("usuario_creo_orden"),
            columnas.pop(columnas.index("supervisor")),
        )
        df_combinado = df_combinado[columnas]
        df_combinado.drop("codigo", axis=1, inplace=True)

        # Asegurar compatibilidad de tipos para el merge con incentivos
        df_combinado["Grupo"] = df_combinado["Grupo"].astype(str)
        df_combinado["Grupo_Anterior"] = df_combinado["Grupo_Anterior"].astype(str)
        df_combinado["tipo_venta"] = df_combinado["tipo_venta"].astype(str)

        data_incentivos["Grupo"] = data_incentivos["Grupo"].astype(str)
        data_incentivos["Grupo_Anterior"] = data_incentivos["Grupo_Anterior"].astype(str)
        data_incentivos["tipo_venta"] = data_incentivos["tipo_venta"].astype(str)

        # depuración para verificar claves de unión
        """
        print("Valores únicos en df_combinado:")
        print(f"tipo_venta: {df_combinado['tipo_venta'].unique()}")
        print(f"Grupo: {df_combinado['Grupo'].unique()}")
        print(f"Grupo_Anterior: {df_combinado['Grupo_Anterior'].unique()}")

        print("\nValores únicos en data_incentivos:")
        print(f"tipo_venta: {data_incentivos['tipo_venta'].unique()}")
        print(f"Grupo: {data_incentivos['Grupo'].unique()}")
        print(f"Grupo_Anterior: {data_incentivos['Grupo_Anterior'].unique()}")
        """

        # Realizar la unión con tabla de incentivos
        df_combinado = pd.merge(
            df_combinado,
            data_incentivos,
            on=["tipo_venta", "Grupo", "Grupo_Anterior"],
            how="left"
        )

        # Verificar registros sin coincidencia
        sin_coincidencia = df_combinado[df_combinado["Comision_100"].isna()]
        if not sin_coincidencia.empty:
            print(f"Hay {len(sin_coincidencia)} registros sin coincidencia en incentivos")
            print(sin_coincidencia[["tipo_venta", "Grupo", "Grupo_Anterior"]].head())

        # Convertir comisiones a numérico y rellenar valores nulos
        df_combinado["Comision_100"] = pd.to_numeric(df_combinado["Comision_100"], errors='coerce').fillna(0)
        df_combinado["Comision_75"] = pd.to_numeric(df_combinado["Comision_75"], errors='coerce').fillna(0)

        # Asegurar que total_ventas es numérico
        df_combinado["total_ventas"] = pd.to_numeric(df_combinado["total_ventas"], errors="coerce").fillna(0)

        # Calcular comisiones
        df_combinado["Comision_100"] = df_combinado["total_ventas"] * df_combinado["Comision_100"]
        df_combinado["Comision_75"] = df_combinado["total_ventas"] * df_combinado["Comision_75"]

        # Verificar cálculos
        print("\nEjemplos de cálculos de comisión:")
        muestra = df_combinado[df_combinado["Comision_100"] > 0].head(3)
        for _, row in muestra.iterrows():
            print(f"Venta: {row['total_ventas']}, Comisión 100%: {row['Comision_100']}")

        # Guardar en base de datos
        guardar_ventas_detalle_en_db(df_combinado)           
        print("Ventas generadas y guardadas en la base de datos correctamente. 2")
        return jsonify({"status": "success", "message": "Ventas generadas y guardadas en la base de datos correctamente."})       
        
    except Exception as e:
        print(f"Ocurrió un error: {e}")  
        # Fin agugacion ventas